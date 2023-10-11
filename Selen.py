import pyautogui as pya
import openpyxl as xl
import time
import threading

global child
global parent
global region

def init():
    global findall
    global searchfield
    global findinhy
    global edit
    global cut
    global paste

    #look for the findall, edit, searchfield, and the find button
    print('Analyzing screen...')
    region=(0, 0, 1920, 1080)
    edit=None
    while edit is None:
        edit=pya.locateOnScreen(r'GCOH resources\edit.png',grayscale=True,region=region,confidence=.70)
    print('edit is found in ' + str(pya.center(edit)))
    p,m=pya.center(edit)
    m=m+29
    cut=(p,m)
    n=m+55
    paste=(p,n)
    print('cut is found in ' + str(cut))
    print('paste is found in ' + str(paste))

    findall=None
    while findall is None:
        findall=pya.locateOnScreen(r'GCOH resources\findall.png',grayscale=True,region=region,confidence=.70)
    print('find all is found in ' + str(pya.center(findall)))
    findinhy=None
    while findinhy is None:
        findinhy=pya.locateOnScreen(r'GCOH resources\FindInHierarchy.png',grayscale=True,region=region,confidence=.70)
    s,t=pya.center(findinhy)
    s=s+29
    t=t+29
    searchfield=(s,t)
    print('search field is found in ' + str(searchfield))
    print('ready for further operations.....')



def job(char):
    #goes after init, actual operations like typing in child/parent, will rerun init if need be during cases like errorcheck true
    pya.doubleClick(searchfield)
    pya.typewrite(char)
    pya.click(findall)
    print('Looking for object.....')


def errorcheck():
    #look for message at the top, if exists rerun init
    global error
    end_time=time.time()+1.5
    err=None
    while err is None:
        err=pya.locateOnScreen(r'GCOH resources/error.png',grayscale=False,region=(0, 0, 1920, 1080))

        current_time=time.time()
        if current_time >=end_time:
            error=False
            print("NO errors found, proceeding")
            break
        else:
            error=True
            print("Error encountered in this object")


def finding():
    #actions when findall is clicked
    j = None
    while j is None:
            j = pya.locateOnScreen(r'GCOH resources/FIND SHIT.png', grayscale=False, region=(0, 0, 1920, 1080))
    x, y = pya.center(j)
    x = x + 25
    y = y + 20
    pya.click(x, y)
    print('Object found on' )
    print(x, y)
    print('Continuing operations...')

def cute():
    #actual operations when finding is true
    pya.click(edit)
    time.sleep(.5)
    pya.click(cut)


def pastee():
    #actual operations upon paste
    pya.click(edit)
    time.sleep(.5)
    pya.click(paste)

def main():
    #mainthread
    #final function it works
    wb=xl.load_workbook(filename=r'GCOH resources/GCOHPost.xlsx')
    ws=wb["MDGF"]
    lrow=len(ws['A'])

    i=2

    while i < lrow+1:


        child=ws['A' + str(i)].value
        parent=ws['B' + str(i)].value
        print(child + " goes to " + parent)

        i=i+1
        print(i)

        init()
        job(child)
        errorcheck()

        if error is True:
            print(r"I wont continue anymore")
            feundfind()
            pya.click(findreset)
            print('resetting...')
            time.sleep(1)
            print('reset success... moving on to next row...')
            feundfind()
            pya.click(findreset)
            continue
        else:
            finding()
            time.sleep(1)
            cute()
            time.sleep(1)
            job(parent)
            errorcheck()
            if error is True:
                print(r"I wont continue anymore")
                feundfind()
                pya.doubleClick(findreset)
                pya.click(findreset)
                continue
            else:
                finding()
                time.sleep(1)
                pastee()



def test():
#no loop here for test
    init()
    job('DE50R07490')
    errorcheck()
    if error is True:
        print(r"I wont continue anymore")
        feundfind()
        pya.click(findreset)
        print('resetting...')
        time.sleep(1)
        print('reset success... moving on to next row...')
        feundfind()
        pya.click(findreset)
        #continue
    else:
        finding()
        time.sleep(1)
        cute()
        time.sleep(1)
        job('H000000891')
        errorcheck()
        if error is True:
            print(r"I wont continue anymore")
            feundfind()
            pya.doubleClick(findreset)
            pya.click(findreset)
            #continue
        else:
            finding()
            time.sleep(1)
            pastee()

    print("Done")



def feundfind():
    global findreset
    region=(0, 0, 1920, 1080)
    findreset=None
    while findreset is None:
        findreset=pya.locateCenterOnScreen(r'GCOH resources\feund.png',grayscale=True,region=region,confidence=.70)
    print('find button is found in ' + str(findreset))


main()